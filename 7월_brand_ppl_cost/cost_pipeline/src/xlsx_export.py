"""brand_ppl_summary.csv → 브랜드별 예상지출_{label}.xlsx

기존 '브랜드별 예상지출.xlsx' 템플릿을 복사해서 값만 덮어쓰는 방식은 브랜드 수가
바뀌면(추가/삭제) 셀이 비거나 브랜드가 통째로 빠지는 사고가 나기 쉬워서(실제로 한 번
겪음), 매번 브랜드 리스트 길이에 맞춰 레이아웃을 새로 그리는 방식으로 만든다.
"""
import csv
import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

NUM_FMT = "#,##0_ "
BLOCK_COLS = ["brand", "ig_ppl", "yt_ppl", "ig_views", "yt_views", "ig_cost", "yt_cost"]


def _write_block(ws, start_col: int, rows: list):
    c0, c1, c2, c3, c4, c5, c6 = (get_column_letter(start_col + i) for i in range(7))

    ws.merge_cells(f"{c1}1:{c2}1")
    ws.merge_cells(f"{c3}1:{c4}1")
    ws.merge_cells(f"{c5}1:{c6}1")
    ws[f"{c1}1"] = "PPL 수"
    ws[f"{c3}1"] = "합산 조회수"
    ws[f"{c5}1"] = "예상 지출"
    for col in (c1, c3, c5):
        ws[f"{col}1"].alignment = Alignment(horizontal="center")
        ws[f"{col}1"].font = Font(bold=True)

    for col in (c1, c2, c3, c4, c5, c6):
        pass
    ws[f"{c1}2"], ws[f"{c2}2"] = "인스타그램", "유튜브"
    ws[f"{c3}2"], ws[f"{c4}2"] = "인스타그램", "유튜브"
    ws[f"{c5}2"], ws[f"{c6}2"] = "인스타그램", "유튜브"
    for col in (c1, c2, c3, c4, c5, c6):
        ws[f"{col}2"].font = Font(bold=True)

    ws.column_dimensions[c0].width = 14
    for col in (c1, c2, c3, c4, c5, c6):
        ws.column_dimensions[col].width = 13

    for i, row in enumerate(rows):
        r = 3 + i
        ws[f"{c0}{r}"] = row["brand"]
        ws[f"{c1}{r}"] = row["ig_ppl"]
        ws[f"{c2}{r}"] = row["yt_ppl"]
        ws[f"{c3}{r}"] = row["ig_views"]
        ws[f"{c4}{r}"] = row["yt_views"]
        ws[f"{c5}{r}"] = row["ig_cost"]
        ws[f"{c6}{r}"] = row["yt_cost"]
        for col in (c1, c2, c3, c4, c5, c6):
            ws[f"{col}{r}"].number_format = NUM_FMT


def build_xlsx(summary_csv: Path, xlsx_path: Path):
    with open(summary_csv, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        records = [
            {
                "brand": row["brand"],
                "ig_ppl": int(row["인스타 광고 수"]),
                "yt_ppl": int(row["유튜브 광고 수"]),
                "ig_views": int(row["인스타 조회수"]),
                "yt_views": int(row["유튜브 조회수"]),
                "ig_cost": int(row["인스타 비용"]),
                "yt_cost": int(row["유튜브 비용"]),
            }
            for row in reader
        ]

    half = math.ceil(len(records) / 2)
    left, right = records[:half], records[half:]

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    _write_block(ws, start_col=1, rows=left)   # A..G
    if right:
        _write_block(ws, start_col=9, rows=right)  # I..N (H는 비워서 원본과 동일한 여백 유지)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
